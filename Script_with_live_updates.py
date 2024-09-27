import time

import openpyxl
import pandas as pd
import requests
from openpyxl.utils.dataframe import dataframe_to_rows

# Required Libraries
# " pip install gspread oauth2client "
# " pip install pandas openpyxl requests "


# Personal Path For OneDrive Folder
excel_file_path = r'Your path to the OneDrive Folder'

# Fetching Live Cryptocurrency Data
def fetch_crypto_data():
    url = 'https://api.coingecko.com/api/v3/coins/markets'
    params = {
        'vs_currency': 'usd',
        'order': 'market_cap_desc',
        'per_page': 50,
        'page': 1,
        'sparkline': 'false'
    }
    response = requests.get(url, params=params)
    data = response.json()
    
    # Pandas DataFrame with Selected Columns
    df = pd.DataFrame(data, columns=[
        'name', 'symbol', 'current_price', 'market_cap', 'total_volume', 'price_change_percentage_24h'
    ])
    
    return df

# Performing Basic Analysis on Cryptocurrency Data
def analyze_data(df):
    # 1. Identifying the top 5 cryptocurrencies by market capitalization
    top_5_by_market_cap = df.nlargest(5, 'market_cap')

    # 2. Calculating the average price of the top 50 cryptocurrencies
    average_price = df['current_price'].mean()

    # 3. Analyzing the highest and lowest 24-hour percentage price change
    highest_change = df.loc[df['price_change_percentage_24h'].idxmax()]
    lowest_change = df.loc[df['price_change_percentage_24h'].idxmin()]

    analysis_summary = {
        'Top 5 Cryptos by Market Cap': top_5_by_market_cap[['name', 'market_cap']],
        'Average Price': average_price,
        'Highest 24h Change': highest_change[['name', 'price_change_percentage_24h']],
        'Lowest 24h Change': lowest_change[['name', 'price_change_percentage_24h']]
    }

    return analysis_summary

# Updating Data in an Excel Sheet
def update_excel(df, file_path):
    try:
        # Loading Existing Workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        # If no file, create a new workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Crypto Data"
    
    # Clear Existing Data in the Sheet
    for row in sheet['A2:F100']:
        for cell in row:
            cell.value = None
    
    # Writing New Data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
    
    # Save the Updated Excel File
    workbook.save(file_path)
    print("Excel updated at", time.strftime("%Y-%m-%d %H:%M:%S"))

# Infinite loop to fetch data, analyze it, and update Excel every 5 minutes
def main():
    while True:
        # Fetch live cryptocurrency data
        crypto_data = fetch_crypto_data()
        
        # Perform analysis on the data
        analysis = analyze_data(crypto_data)
        
        # Print the analysis results (this can be removed or logged)
        print("\n--- Analysis Summary ---")
        print(f"Top 5 Cryptos by Market Cap:\n{analysis['Top 5 Cryptos by Market Cap']}")
        print(f"Average Price of Top 50 Cryptos: ${analysis['Average Price']:.2f}")
        print(f"Highest 24h Change: {analysis['Highest 24h Change']['name']} ({analysis['Highest 24h Change']['price_change_percentage_24h']:.2f}%)")
        print(f"Lowest 24h Change: {analysis['Lowest 24h Change']['name']} ({analysis['Lowest 24h Change']['price_change_percentage_24h']:.2f}%)\n")
        
        # Update the Excel sheet with new data
        update_excel(crypto_data, excel_file_path)
        
        # Wait for 5 minutes before the next update
        time.sleep(300)

# Run the Script
if __name__ == "__main__":
    main()


##### LINK FOR LIVE UPDATES EXCEL SHEET HAS BEEN INCLUDED IN THE REPORT #####
